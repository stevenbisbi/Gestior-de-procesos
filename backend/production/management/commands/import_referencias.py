"""
Importa el catálogo real de la empresa (Excel de referencias) al sistema:
crea/actualiza TubeSpec (tubos largos) y ProductType (referencias).

Uso:
    python manage.py import_referencias "C:/ruta/Datos Va.xlsx"
    python manage.py import_referencias "C:/ruta/Datos Va.xlsx" --dry-run

Es idempotente: identifica el producto por su item y el tubo largo por su
"Item tubo largo"; si ya existen, los actualiza en lugar de duplicar.

Para importar a PRODUCCIÓN sin instalar nada en el servidor: corré este
comando desde tu máquina apuntando a la base remota:
    set DATABASE_URL=<External Database URL de Render>
    python manage.py import_referencias "C:/ruta/Datos Va.xlsx"

Columnas que usa (los nombres se normalizan, mayúsculas/acentos no importan):
    ítem, Descripción, Longitud [mm], Tolerancia [mm], Espesor [mm],
    Tipo de Lam, Item tubo largo, Descripción tubo largo, Empaque, Cliente,
    Pieza, Proyecto, PROCESO CORTE/MOLETEADO/CHAFLANEADO/CURVADO/
    CONIFICADO/TRONZADORA (marcados con X)
"""

import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, IntegrityError

from production.models import TubeSpec, ProductType


def norm(s):
    """minúsculas, sin acentos, espacios colapsados — para casar encabezados."""
    s = str(s or '').strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s)


# encabezado normalizado → clave interna
HEADERS = {
    'item':                   'item',
    'descripcion':            'descripcion',
    'longitud [mm]':          'longitud',
    'tolerancia [mm]':        'tolerancia',
    'espesor [mm]':           'espesor',
    'tipo de lam':            'tipo_lam',
    'item tubo largo':        'tubo_item',
    'descripcion tubo largo': 'tubo_desc',
    'empaque':                'empaque',
    'cliente':                'cliente',
    'pieza':                  'pieza',
    'proyecto':               'proyecto',
    'proceso corte':          'p_corte',
    'proceso moleteado':      'p_moleteo',
    'proceso chaflaneado':    'p_chaflan',
    'proceso curvado':        'p_curvado',
    'proceso conificado':     'p_conificado',
    'proceso tronzadora':     'p_tronzadora',
}

# fracciones tipo "1 1/2" o "3/4", o decimales "22.2"
NUM = r'(\d+\s+\d+/\d+|\d+/\d+|\d+(?:[.,]\d+)?)'
RE_TRIPLE = re.compile(NUM + r'\s*[x*]\s*' + NUM + r'\s*[x*]\s*' + NUM, re.I)
RE_NUMS   = re.compile(NUM)


def to_float(s):
    s = str(s).strip().replace(',', '.')
    if '/' in s:  # fracción → no es convertible; se usa como texto (diámetro)
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_material(*texts):
    """Detecta el material en descripciones/columna Tipo de Lam."""
    joined = ' ' + norm(' '.join(str(t) for t in texts if t)) + ' '
    if 'cr-est' in joined or 'cr est' in joined:  return 'cr_est'
    if 'hr-est' in joined or 'hr est' in joined:  return 'hr_est'
    if 'galv' in joined or ' gv ' in joined:      return 'gv'
    if ' ec ' in joined or 'ec spfc' in joined:   return 'ec'
    if ' hr ' in joined:                          return 'hr'
    if ' cr ' in joined:                          return 'cr'
    return None


def parse_tube(desc, espesor_col):
    """
    Extrae (diámetro, espesor, longitud) de la descripción del tubo largo.
    Formatos reales: 'TUBO CR EST 22.2x2.0x6349 mm',
    'TUBO 22.2 CAL.2.0 (CR-EST) * 5985 mm', 'TUB (CR-EST) REDONDO 22.2x2.0x6048mm'
    """
    desc = str(desc or '')
    m = RE_TRIPLE.search(desc)
    if m:
        od, th, ln = m.group(1), to_float(m.group(2)), to_float(m.group(3))
        return od.strip(), th, ln
    # Sin patrón AxBxC: buscar números sueltos
    nums = RE_NUMS.findall(desc)
    length = None
    od = None
    th = to_float(espesor_col)
    for n in nums:
        f = to_float(n)
        if f is not None and f > 1000 and length is None:
            length = f
    for n in nums:
        f = to_float(n)
        if f is not None and f == length:
            continue
        if th is not None and f == th:
            continue
        od = n.strip()
        break
    return od, th, length


class Command(BaseCommand):
    help = 'Importa referencias y tubos largos desde el Excel de la empresa.'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta del .xlsx')
        parser.add_argument('--dry-run', action='store_true',
                            help='Simula: muestra lo que haría sin guardar nada.')
        parser.add_argument('--hoja', default=None, help='Nombre de la hoja (default: la primera)')

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('Falta openpyxl. Instalalo con: pip install openpyxl')

        wb = load_workbook(opts['archivo'], data_only=True, read_only=True)
        ws = wb[opts['hoja']] if opts['hoja'] else wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows)
        col_idx = {}   # clave interna → índice de columna
        for i, h in enumerate(header_row):
            key = HEADERS.get(norm(h))
            if key and key not in col_idx:
                col_idx[key] = i

        missing = [k for k in ('item', 'descripcion', 'longitud') if k not in col_idx]
        if missing:
            raise CommandError(f'No encontré columnas obligatorias: {missing}. '
                               f'Encabezados vistos: {[norm(h) for h in header_row if h]}')

        def val(row, key):
            i = col_idx.get(key)
            v = row[i] if i is not None and i < len(row) else None
            return None if v is None or (isinstance(v, str) and not v.strip()) else v

        def marked(row, key):
            v = val(row, key)
            return bool(v) and str(v).strip().upper() in ('X', 'SI', 'SÍ', '1', 'TRUE')

        stats = {'tubos_nuevos': 0, 'tubos_upd': 0, 'prod_nuevos': 0, 'prod_upd': 0, 'saltadas': 0}
        warnings = []

        with transaction.atomic():
            for n, row in enumerate(rows, start=2):   # 2 = primera fila de datos en Excel
                item = val(row, 'item')
                desc = val(row, 'descripcion')
                if not item or not desc:
                    stats['saltadas'] += 1
                    continue
                item = str(item).strip().rstrip('.0') if isinstance(item, float) else str(item).strip()

                cut_length = to_float(val(row, 'longitud'))
                if not cut_length:
                    warnings.append(f'Fila {n} (item {item}): sin longitud de corte — saltada.')
                    stats['saltadas'] += 1
                    continue

                # ── Tubo largo ────────────────────────────────────────
                tubo_item = val(row, 'tubo_item')
                tubo_item = str(int(tubo_item)) if isinstance(tubo_item, float) else str(tubo_item or '').strip()
                tubo_desc = val(row, 'tubo_desc')
                od, th, largo = parse_tube(tubo_desc, val(row, 'espesor'))
                material = parse_material(tubo_desc, desc, val(row, 'tipo_lam')) or 'cr'
                shape = 'square' if re.search(r'rect|cuad', norm(tubo_desc) + norm(desc)) else 'round'
                if not od or not th or not largo:
                    warnings.append(f'Fila {n} (item {item}): no pude leer el tubo largo de '
                                    f'"{tubo_desc}" (Ø={od}, esp={th}, largo={largo}) — saltada.')
                    stats['saltadas'] += 1
                    continue

                tube_fields = {'shape': shape, 'outer_diameter': od, 'thickness': th,
                               'material': material, 'original_length': largo}
                try:
                    if tubo_item:
                        tube, created = TubeSpec.objects.update_or_create(
                            item_code=tubo_item, defaults=tube_fields)
                    else:
                        tube, created = TubeSpec.objects.get_or_create(
                            outer_diameter=od, thickness=th, material=material,
                            original_length=largo, defaults={'shape': shape})
                except IntegrityError:
                    # Ya existe un tubo con esas dimensiones bajo otro item — se reutiliza
                    tube = TubeSpec.objects.filter(
                        outer_diameter=od, thickness=th, material=material,
                        original_length=largo).first()
                    created = False
                    warnings.append(f'Fila {n}: tubo {tubo_item} tiene las mismas dimensiones '
                                    f'que el item {tube.item_code or "(sin item)"} — se reutilizó.')
                stats['tubos_nuevos' if created else 'tubos_upd'] += 1

                # ── Procesos ──────────────────────────────────────────
                for proc, label in (('p_conificado', 'CONIFICADO'), ('p_tronzadora', 'TRONZADORA')):
                    if marked(row, proc):
                        warnings.append(f'Fila {n} (item {item}): proceso {label} marcado — '
                                        f'el sistema no lo maneja, queda solo en notas.')

                # ── Notas: lo que el sistema no modela como campo ─────
                notes = []
                if val(row, 'pieza'):       notes.append(f"Pieza: {str(val(row,'pieza')).strip()}")
                if val(row, 'tolerancia') is not None:
                    notes.append(f"Tolerancia: ±{val(row,'tolerancia')} mm")
                if val(row, 'proyecto'):    notes.append(f"Proyecto: {str(val(row,'proyecto')).strip()}")
                if marked(row, 'p_conificado'): notes.append('Requiere CONIFICADO (fuera del sistema)')
                if marked(row, 'p_tronzadora'): notes.append('Requiere TRONZADORA (fuera del sistema)')

                # ── Producto ──────────────────────────────────────────
                prod_fields = {
                    'name':             str(desc).strip(),
                    'tube_spec':        tube,
                    'cut_length':       cut_length,
                    'client':           str(val(row, 'cliente') or '').strip(),
                    'packaging':        str(val(row, 'empaque') or '').strip(),
                    'requires_chaflan': marked(row, 'p_chaflan'),
                    'requires_moleteo': marked(row, 'p_moleteo'),
                    'requires_curvado': marked(row, 'p_curvado'),
                    'notes':            ' | '.join(notes),
                }
                _, created = ProductType.objects.update_or_create(
                    item_code=item, defaults=prod_fields)
                stats['prod_nuevos' if created else 'prod_upd'] += 1

            if opts['dry_run']:
                transaction.set_rollback(True)

        # ── Reporte ───────────────────────────────────────────────────
        modo = 'SIMULACIÓN (nada se guardó)' if opts['dry_run'] else 'Importación completada'
        self.stdout.write(self.style.SUCCESS(
            f"\n{modo}\n"
            f"  Tubos largos:  {stats['tubos_nuevos']} nuevos, {stats['tubos_upd']} actualizados\n"
            f"  Referencias:   {stats['prod_nuevos']} nuevas, {stats['prod_upd']} actualizadas\n"
            f"  Filas saltadas: {stats['saltadas']}"))
        if warnings:
            self.stdout.write(self.style.WARNING(f'\nAvisos ({len(warnings)}):'))
            for w in warnings:
                self.stdout.write(self.style.WARNING(f'  - {w}'))
